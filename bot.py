"""
Telegram-бот фиксации задач строительного предприятия
"""

import logging
import os
import sqlite3
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
    Application, CallbackQueryHandler, CommandHandler,
    ContextTypes, ConversationHandler, MessageHandler, filters,
)

from templates_module import (
    get_template_handlers,
    get_admin_template_handlers,
    init_templates_db,
)

# Конфиг
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
ADMIN_ID  = int(os.getenv("ADMIN_ID", "0"))

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    level=logging.INFO,
)
log = logging.getLogger(__name__)

BASE_DIR   = Path(__file__).parent
DATA_DIR   = BASE_DIR / "data"
EXPORT_DIR = DATA_DIR / "exports"
DB_PATH    = DATA_DIR / "tasks.db"
DATA_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)


# БАЗА ДАННЫХ
def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS objects (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            address    TEXT DEFAULT '',
            obj_type   TEXT DEFAULT '',
            status     TEXT DEFAULT 'Активный',
            created_by TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS bp (
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            name   TEXT NOT NULL UNIQUE,
            active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS tasks_ref (
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            bp_id  INTEGER NOT NULL REFERENCES bp(id),
            name   TEXT NOT NULL,
            active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS procedures (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL REFERENCES tasks_ref(id),
            name    TEXT NOT NULL,
            active  INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS work_log (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            date           TEXT NOT NULL,
            employee_name  TEXT NOT NULL,
            employee_tg_id TEXT NOT NULL,
            object_id      INTEGER,
            object_name    TEXT NOT NULL,
            bp_id          INTEGER,
            bp_name        TEXT NOT NULL,
            task_id        INTEGER,
            task_name      TEXT NOT NULL,
            procedure_id   INTEGER,
            procedure_name TEXT NOT NULL,
            time_start     TEXT NOT NULL,
            time_end       TEXT NOT NULL,
            duration_h     REAL DEFAULT 0,
            coworkers      TEXT DEFAULT '',
            description    TEXT DEFAULT '',
            created_at     TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS planned_tasks (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            title          TEXT NOT NULL,
            object_name    TEXT DEFAULT '',
            assignee_name  TEXT NOT NULL,
            assignee_tg_id TEXT DEFAULT '',
            planned_date   TEXT NOT NULL,
            planned_time   TEXT DEFAULT '',
            status         TEXT DEFAULT 'Запланирована',
            day_status     TEXT DEFAULT '',
            consistency    TEXT DEFAULT 'Согласована',
            created_by     TEXT DEFAULT '',
            created_at     TEXT DEFAULT '',
            reminded_2h    INTEGER DEFAULT 0,
            reminded_day   INTEGER DEFAULT 0,
            bp_name        TEXT DEFAULT '',
            task_ref_name  TEXT DEFAULT '',
            fail_reason    TEXT DEFAULT '',
            rescheduled_to TEXT DEFAULT ''
        );
        """)
        if not db.execute("SELECT 1 FROM bp LIMIT 1").fetchone():
            _seed(db)
        db.commit()
        init_templates_db(db)


def _seed(db):
    tree = {
        "Монтаж системы отопления": {
            "Монтаж котла":               ["Обвязка котла", "Подключение газа/электро", "Настройка котла"],
            "Монтаж насосной группы":     ["Установка насоса", "Подключение трубопровода"],
            "Монтаж трубопровода":        ["Прокладка труб", "Пайка/пресс фитинги", "Опрессовка системы"],
            "Монтаж радиаторов":          ["Установка кронштейнов", "Навеска радиатора", "Подключение"],
            "Сборка котельной под ключ":  ["Монтаж котла", "Монтаж гидрострелки", "Монтаж коллектора", "Пусконаладка"],
            "Заполнение и пусконаладка":  ["Заполнение теплоносителем", "Проверка герметичности", "Настройка давления"],
        },
        "Монтаж системы снеготаяния": {
            "Укладка греющего кабеля":  ["Разметка", "Укладка кабеля", "Фиксация кабеля"],
            "Монтаж терморегулятора":   ["Установка терморегулятора", "Подключение к щиту"],
            "Пусконаладка":             ["Тест системы", "Настройка терморегулятора"],
        },
        "Монтаж автоматизации котельного оборудования": {
            "Монтаж контроллера":  ["Установка контроллера", "Подключение датчиков температуры"],
            "Программирование":    ["Настройка алгоритма", "Тест автоматики"],
            "Монтаж GSM-модуля":   ["Установка SIM-карты", "Настройка уведомлений"],
        },
        "Монтаж радиаторов": {
            "Монтаж радиаторов": ["Установка кронштейнов", "Навеска радиатора", "Подключение"],
        },
        "Монтаж теплого пола": {
            "Укладка труб": ["Разметка контуров", "Укладка трубы", "Опрессовка контуров"],
            "Монтаж коллектора": ["Установка коллектора", "Подключение контуров"],
        },
        "Монтаж. Прочее": {
            "Прочий монтаж": ["Работа по заданию (описать)"],
        },
        "Обслуживание системы отопления": {
            "Плановое ТО":          ["Чистка фильтров", "Проверка давления", "Проверка теплоносителя"],
            "Исправление протечки": ["Поиск протечки", "Устранение протечки", "Опрессовка после ремонта"],
            "Опрессовка системы":   ["Подготовка к опрессовке", "Подача давления", "Оформление акта"],
        },
        "Обслуживание котельной": {
            "Плановое ТО котла": ["Чистка теплообменника", "Чистка горелки", "Замена фильтров", "Проверка тяги"],
            "Ремонт котла":      ["Диагностика неисправности", "Замена запчасти", "Тест после ремонта"],
            "Запуск котельной":  ["Первичный запуск", "Настройка параметров", "Инструктаж клиента"],
        },
        "Аварийное обслуживание": {
            "Аварийный выезд": ["Диагностика аварии", "Устранение аварии", "Фиксация в журнале"],
        },
        "Автоматизация": {
            "Монтаж автоматики": ["Установка оборудования", "Настройка контроллера"],
        },
        "Уличные работы": {
            "Земляные работы": ["Рытьё траншеи", "Укладка трубы в траншею", "Засыпка траншеи"],
            "Мощение":         ["Укладка плитки", "Установка бордюра"],
        },
        "Логистика": {
            "Поездка":            ["Поездка на объект", "Поездка за материалами", "Поездка к поставщику"],
            "Разгрузка погрузка": ["Разгрузка материалов на объекте", "Погрузка инструментов"],
        },
        "Документооборот": {
            "Договоры": ["Составление договора", "Согласование с клиентом", "Подписание договора"],
            "Сметы":    ["Составление сметы", "Согласование сметы", "Внесение правок"],
            "Акты":     ["Составление акта выполненных работ", "Подписание акта", "Отправка акта клиенту"],
        },
        "Встреча": {
            "Встреча с клиентом новый объект":      ["Первичный осмотр объекта", "Обсуждение ТЗ", "Повторный выезд с мастерами"],
            "Встреча с клиентом действующий объект":["Осмотр хода работ", "Согласование изменений", "Приёмка работ клиентом"],
            "Внутренняя встреча":                   ["Планёрка команды", "Разбор задач по объектам"],
        },
        "Оргуправление": {
            "Закупки":                 ["Подбор материалов", "Заказ у поставщика", "Оплата счёта"],
            "Административные задачи": ["Подача документов", "Работа с документами", "Прочее"],
        },
        "Исправление брака": {
            "Переделка": ["Демонтаж неправильного", "Повторный монтаж", "Проверка после исправления"],
        },
    }
    for bp_name, tasks in tree.items():
        db.execute("INSERT INTO bp (name) VALUES (?)", (bp_name,))
        bp_id = db.execute("SELECT id FROM bp WHERE name=?", (bp_name,)).fetchone()["id"]
        for task_name, procs in tasks.items():
            db.execute("INSERT INTO tasks_ref (bp_id, name) VALUES (?,?)", (bp_id, task_name))
            task_id = db.execute(
                "SELECT id FROM tasks_ref WHERE bp_id=? AND name=?", (bp_id, task_name)
            ).fetchone()["id"]
            for proc in procs:
                db.execute("INSERT INTO procedures (task_id, name) VALUES (?,?)", (task_id, proc))
    for name, addr, otype in [
        ("Офис компании", "г. Ульяновск, ул. Главная, 1", "Офис"),
        ("Склад",         "г. Ульяновск, ул. Складская, 5", "Склад"),
    ]:
        db.execute(
            "INSERT INTO objects (name, address, obj_type, created_by, created_at) VALUES (?,?,?,?,?)",
            (name, addr, otype, "Система", datetime.now().strftime("%d.%m.%Y"))
        )


# EXCEL ЭКСПОРТ
def export_to_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    hf  = PatternFill("solid", fgColor="4472C4")
    hft = Font(bold=True, color="FFFFFF", size=10)
    af  = PatternFill("solid", fgColor="DCE6F1")
    t   = Side(style="thin", color="AAAAAA")
    brd = Border(left=t, right=t, top=t, bottom=t)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "Задачи"
    h1 = ["ID","Дата","Сотрудник","TG_ID","Объект","Бизнес-процесс","Задача","Процедура",
          "Начало","Конец","Длит. ч","Коллеги","Описание","Создано"]
    w1 = [5,11,18,12,22,28,28,28,8,8,9,20,30,18]
    for c,(h,w) in enumerate(zip(h1,w1),1):
        cl = ws1.cell(row=1, column=c, value=h)
        cl.fill=hf; cl.font=hft; cl.alignment=ctr; cl.border=brd
        ws1.column_dimensions[cl.column_letter].width=w
    ws1.row_dimensions[1].height=22
    ws1.freeze_panes="A2"

    with get_db() as db:
        rows = db.execute("""SELECT id,date,employee_name,employee_tg_id,
            object_name,bp_name,task_name,procedure_name,
            time_start,time_end,duration_h,coworkers,description,created_at
            FROM work_log ORDER BY id""").fetchall()
    for r, row in enumerate(rows, 2):
        fill = af if r % 2 == 0 else None
        for c, val in enumerate(row, 1):
            cl = ws1.cell(row=r, column=c, value=val)
            cl.border = brd
            cl.alignment = ctr if c in (1,9,10,11) else lft
            if fill:
                cl.fill = fill

    ws2 = wb.create_sheet("Объекты")
    for c,(h,w) in enumerate(zip(
        ["ID","Наименование","Адрес","Тип объекта","Статус","Добавил","Дата добавления"],
        [5,25,35,18,14,18,18]
    ),1):
        cl = ws2.cell(row=1, column=c, value=h)
        cl.fill=hf; cl.font=hft; cl.alignment=ctr; cl.border=brd
        ws2.column_dimensions[cl.column_letter].width=w
    with get_db() as db:
        objs = db.execute("SELECT id,name,address,obj_type,status,created_by,created_at FROM objects").fetchall()
    for r, obj in enumerate(objs, 2):
        for c, val in enumerate(obj, 1):
            cl = ws2.cell(row=r, column=c, value=val)
            cl.border=brd; cl.alignment=ctr if c==1 else lft

    ws3 = wb.create_sheet("Справочник_БП")
    for c,(h,w) in enumerate(zip(["БП","Задача","Процедура"],[30,30,35]),1):
        cl = ws3.cell(row=1, column=c, value=h)
        cl.fill=hf; cl.font=hft; cl.alignment=ctr; cl.border=brd
        ws3.column_dimensions[cl.column_letter].width=w
    with get_db() as db:
        ref = db.execute("""SELECT b.name,t.name,p.name FROM bp b
            JOIN tasks_ref t ON t.bp_id=b.id
            JOIN procedures p ON p.task_id=t.id
            WHERE b.active=1 AND t.active=1 AND p.active=1
            ORDER BY b.name,t.name,p.name""").fetchall()
    for r, row in enumerate(ref, 2):
        for c, val in enumerate(row, 1):
            cl = ws3.cell(row=r, column=c, value=val)
            cl.border=brd; cl.alignment=lft

    path = EXPORT_DIR / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    return path


# СОСТОЯНИЯ
(U_OBJECT, U_OBJ_NEW_NAME, U_OBJ_NEW_ADDR, U_OBJ_NEW_TYPE,
 U_BP, U_TASK, U_TASK_CUSTOM, U_PROCEDURE, U_PROC_CUSTOM,
 U_TIME_START, U_TIME_END, U_COWORKERS, U_CW_INPUT,
 U_DESCRIPTION, U_CONFIRM) = range(15)

(A_MENU, A_OBJ_NAME, A_OBJ_ADDR, A_OBJ_TYPE,
 A_BP_NAME, A_BP_TASK_NAME, A_SEL_BP, A_TASK_NAME,
 A_SEL_TASK, A_PROC_NAME) = range(100, 110)

# Состояния диалога планировщика
P_DATE, P_TITLE, P_OBJECT, P_BP, P_TASK, P_ASSIGNEE, P_TIME_START, P_TIME_END, P_CONFIRM = range(200, 209)
# Состояния обработки невыполненной задачи
PS_FAIL_REASON, PS_RESCHEDULE_DATE = range(300, 302)


# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
def is_admin(uid: int) -> bool:
    return uid == ADMIN_ID


def _kb(items, cols=2, prefix="") -> InlineKeyboardMarkup:
    btns = [InlineKeyboardButton(str(i), callback_data=f"{prefix}{i}") for i in items]
    rows = [btns[i:i+cols] for i in range(0, len(btns), cols)]
    return InlineKeyboardMarkup(rows)


def obj_types_kb() -> InlineKeyboardMarkup:
    """Шаг 1: выбор ТИПА объекта."""
    with get_db() as db:
        types = db.execute(
            "SELECT DISTINCT obj_type FROM objects WHERE obj_type != '' ORDER BY obj_type"
        ).fetchall()
    rows = [[InlineKeyboardButton(f"🏗 {t['obj_type']}", callback_data=f"OTYPE_{t['obj_type']}")] for t in types]
    rows.append([InlineKeyboardButton("➕ Новый объект", callback_data="OBJ_NEW")])
    return InlineKeyboardMarkup(rows)


def objects_by_type_kb(obj_type: str) -> InlineKeyboardMarkup:
    """Шаг 2: список объектов выбранного типа."""
    with get_db() as db:
        objs = db.execute(
            "SELECT id, name, address FROM objects WHERE obj_type=? ORDER BY name",
            (obj_type,)
        ).fetchall()
    rows = []
    for obj in objs:
        lbl = obj["name"]
        if obj["address"]:
            lbl += f" ({obj['address'][:22]})"
        rows.append([InlineKeyboardButton(f"📍 {lbl}", callback_data=f"OBJ_{obj['id']}")])
    rows.append([InlineKeyboardButton("◀️ Назад к типам", callback_data="OBJ_BACK_TYPES")])
    return InlineKeyboardMarkup(rows)


def objects_kb() -> InlineKeyboardMarkup:
    """Fallback: все объекты без фильтра (используется в /plan)."""
    with get_db() as db:
        objs = db.execute("SELECT id, name, address FROM objects ORDER BY name").fetchall()
    rows = []
    for obj in objs:
        lbl = obj["name"]
        if obj["address"]:
            lbl += f" ({obj['address'][:22]})"
        rows.append([InlineKeyboardButton(f"📍 {lbl}", callback_data=f"OBJ_{obj['id']}")])
    rows.append([InlineKeyboardButton("➕ Новый объект", callback_data="OBJ_NEW")])
    return InlineKeyboardMarkup(rows)


def bp_kb() -> InlineKeyboardMarkup:
    with get_db() as db:
        bps = db.execute("SELECT id, name FROM bp WHERE active=1 ORDER BY name").fetchall()
    rows = []
    for i in range(0, len(bps), 2):
        row = [InlineKeyboardButton(b["name"], callback_data=f"BP_{b['id']}") for b in bps[i:i+2]]
        rows.append(row)
    return InlineKeyboardMarkup(rows)


def tasks_kb(bp_id) -> InlineKeyboardMarkup:
    with get_db() as db:
        ts = db.execute(
            "SELECT id, name FROM tasks_ref WHERE bp_id=? AND active=1 ORDER BY name",
            (bp_id,)
        ).fetchall()
    rows = [[InlineKeyboardButton(t["name"], callback_data=f"TASK_{t['id']}")] for t in ts]
    rows.append([InlineKeyboardButton("✏️ Своя задача", callback_data="TASK_CUSTOM")])
    return InlineKeyboardMarkup(rows)


def procs_kb(task_id) -> InlineKeyboardMarkup:
    with get_db() as db:
        ps = db.execute(
            "SELECT id, name FROM procedures WHERE task_id=? AND active=1 ORDER BY name",
            (task_id,)
        ).fetchall()
    rows = [[InlineKeyboardButton(p["name"], callback_data=f"PROC_{p['id']}")] for p in ps]
    rows.append([InlineKeyboardButton("✏️ Своя процедура", callback_data="PROC_CUSTOM")])
    return InlineKeyboardMarkup(rows)


def cw_kb() -> InlineKeyboardMarkup:
    with get_db() as db:
        ns = db.execute(
            "SELECT DISTINCT employee_name FROM work_log ORDER BY employee_name LIMIT 10"
        ).fetchall()
    rows = [[InlineKeyboardButton(n["employee_name"], callback_data=f"CW_{n['employee_name']}")] for n in ns]
    rows.append([
        InlineKeyboardButton("✏️ Ввести имя", callback_data="CW_CUSTOM"),
        InlineKeyboardButton("👤 Один",       callback_data="CW_SOLO"),
    ])
    return InlineKeyboardMarkup(rows)


def confirm_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Сохранить", callback_data="CYES"),
        InlineKeyboardButton("❌ Отменить",  callback_data="CNO"),
    ]])


def dur_h(start, end) -> float:
    try:
        ts = datetime.strptime(start, "%H:%M")
        te = datetime.strptime(end,   "%H:%M")
        d  = (te - ts).seconds / 3600
        return round(d if d >= 0 else d + 24, 2)
    except Exception:
        return 0.0


def make_summary(d: dict) -> str:
    h    = dur_h(d.get("time_start",""), d.get("time_end",""))
    cw   = d.get("coworkers","")
    team = f"+ {cw}" if cw else "один"
    desc = d.get("description","")
    lines = [
        "Проверь запись перед сохранением:",
        "",
        f"Сотрудник : {d.get('employee_name','')}",
        f"Дата      : {d.get('date','')}",
        f"Время     : {d.get('time_start','')} - {d.get('time_end','')}  ({h} ч)",
        f"Объект    : {d.get('object_name','')}",
        f"БП        : {d.get('bp_name','')}",
        f"Задача    : {d.get('task_name','')}",
        f"Процедура : {d.get('proc_name','')}",
        f"Бригада   : {team}",
    ]
    if desc:
        lines.append(f"Описание  : {desc}")
    return "\n".join(lines)


def save_work(d: dict) -> int:
    with get_db() as db:
        db.execute("""
            INSERT INTO work_log
                (date, employee_name, employee_tg_id,
                 object_id, object_name,
                 bp_id, bp_name, task_id, task_name,
                 procedure_id, procedure_name,
                 time_start, time_end, duration_h,
                 coworkers, description, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            d["date"], d["employee_name"], d["employee_tg_id"],
            d.get("object_id"),  d["object_name"],
            d.get("bp_id"),      d["bp_name"],
            d.get("task_id"),    d["task_name"],
            d.get("proc_id"),    d["proc_name"],
            d["time_start"],     d["time_end"],
            d.get("dur_h", 0),
            d.get("coworkers",""), d.get("description",""),
            datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        ))
        db.commit()
        rec_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Закрываем объект если выбрана "Сдача работ"
        if d.get("_close_object") and d.get("object_id"):
            db.execute(
                "UPDATE objects SET status='Сдан' WHERE id=?",
                (d["object_id"],)
            )
            db.commit()
        return rec_id


# ПОЛЬЗОВАТЕЛЬСКИЙ ДИАЛОГ
async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Команда /help — список всех команд."""
    is_adm = is_admin(update.effective_user.id)
    text = (
        "Доступные команды:\n"
        "\n"
        "/start  — зафиксировать выполненную задачу\n"
        "/template  — применить шаблон"
        "/plan   — добавить задачу в расписание\n"
        "/tasks  — задачи на сегодня (с кнопками статуса)\n"
        "/week   — задачи на ближайшие 7 дней\n"
        "/info   — текущий статус дня и активная задача\n"
        "/cancel — отменить текущее действие\n"
        "/help   — эта справка"
    )
    if is_adm:
        text += (
            "\n\n"
            "Команды администратора:\n"
            "/admin  — панель управления (объекты, БП, выгрузка Excel)\n"
            "/templates — управление шаблонами задач\n"
            "/stats  — статистика выполнения задач за сегодня"            
        )
    await update.message.reply_text(text)


async def cmd_info(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    today = datetime.now().strftime("%d.%m.%Y")
    now_t = datetime.now().strftime("%H:%M")
    uid   = str(update.effective_user.id)

    with get_db() as db:
        planned = db.execute("""
            SELECT * FROM planned_tasks
            WHERE planned_date=? AND assignee_tg_id=?
            ORDER BY planned_time, id
        """, (today, uid)).fetchall()

        last_work = db.execute("""
            SELECT * FROM work_log
            WHERE date=? AND employee_tg_id=?
            ORDER BY id DESC LIMIT 1 
        """, (today, uid)).fetchone()

        # Все записи /start за сегодня этого пользователя
        work_today = db.execute("""
            SELECT bp_name, task_name, object_name, time_start, time_end, duration_h
            FROM work_log
            WHERE date=? AND employee_tg_id=?
            ORDER BY id
        """, (today, uid)).fetchall()

    lines = [f"Статус дня ({today})  {now_t}\n"]

    if work_today:
        lines.append(f"Выполненных задач сегодня: {len(work_today)}")
        for w in work_today:
            lines.append(
                f"  ✅ {w['bp_name']} — {w['task_name']}\n"
                f"     {w['object_name']}  {w['time_start']}–{w['time_end']}  ({w['duration_h']} ч)"
            )
        lines.append("")
    else:
        lines.append("Сегодня ещё нет зафиксированных задач (/start).\n")

    if planned:
        icons = {"Запланирована":"🔵","Выполнена":"✅","В процессе":"🔄","Не выполнена":"❌"}
        lines.append("Расписание на сегодня:")
        for t in planned:
            tp   = f" {t['planned_time']}" if t["planned_time"] else ""
            icon = icons.get(t["status"], "•")
            lines.append(f"  {icon} {t['title']}{tp}")
            if t["object_name"]:
                lines.append(f"     Объект: {t['object_name']}")
        done  = sum(1 for t in planned if t["status"] == "Выполнена")
        total = len(planned)
        lines.append(f"\nВыполнено по расписанию: {done} из {total}")
    else:
        lines.append("Плановых задач на сегодня нет.\n/plan — добавить задачу")

    await update.message.reply_text("\n".join(lines))
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    ctx.user_data["date"]           = datetime.now().strftime("%d.%m.%Y")
    ctx.user_data["employee_tg_id"] = str(update.effective_user.id)
    ctx.user_data["employee_name"]  = update.effective_user.full_name or "Сотрудник"

    await update.message.reply_text(
        f"Привет, {ctx.user_data['employee_name']}!\n"
        f"Дата: {ctx.user_data['date']}\n\n"
        "Шаг 1: Выбери тип объекта:",
        reply_markup=obj_types_kb(),
    )
    return U_OBJECT


async def cmd_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text("Отменено. /start — новая запись.")
    return ConversationHandler.END


# Объект 
async def u_object(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    data = q.data

    # Вернуться к выбору типа
    if data == "OBJ_BACK_TYPES":
        await q.edit_message_text(
            "Шаг 1: Выбери тип объекта:",
            reply_markup=obj_types_kb(),
        )
        return U_OBJECT

    # Выбрали тип — показываем объекты этого типа
    if data.startswith("OTYPE_"):
        obj_type = data.replace("OTYPE_", "")
        ctx.user_data["_obj_type_filter"] = obj_type
        await q.edit_message_text(
            f"Тип: {obj_type}\n\nВыбери объект:",
            reply_markup=objects_by_type_kb(obj_type),
        )
        return U_OBJECT

    # Создать новый объект
    if data == "OBJ_NEW":
        await q.edit_message_text(
            "Введи наименование нового объекта\n"
            "Например: Дом Петровых, Котельная ТЦ Восток"
        )
        return U_OBJ_NEW_NAME

    # Выбрали конкретный объект
    obj_id = int(data.split("_")[1])
    with get_db() as db:
        obj = db.execute("SELECT id, name, status FROM objects WHERE id=?", (obj_id,)).fetchone()

    # Проверяем: объект не закрыт?
    if obj["status"] == "Сдан":
        await q.edit_message_text(
            f"Объект {obj['name']} помечен как СДАН.\n"
            "Работы на нём завершены и закрыты.\n\n"
            "Выбери другой объект:",
            reply_markup=objects_by_type_kb(ctx.user_data.get("_obj_type_filter", "")),
        )
        return U_OBJECT

    ctx.user_data["object_id"]   = obj["id"]
    ctx.user_data["object_name"] = obj["name"]

    await q.edit_message_text(
        f"Объект: {obj['name']}\n\nВыбери бизнес-процесс:",
        reply_markup=bp_kb(),
    )
    return U_BP


async def u_obj_new_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("Имя не может быть пустым. Введи ещё раз:")
        return U_OBJ_NEW_NAME
    ctx.user_data["_obj_name"] = name
    await update.message.reply_text("Введи адрес объекта (или - чтобы пропустить):")
    return U_OBJ_NEW_ADDR


async def u_obj_new_addr(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    addr = update.message.text.strip()
    ctx.user_data["_obj_addr"] = "" if addr == "-" else addr
    await update.message.reply_text(
        "Выбери тип объекта:",
        reply_markup=_kb(
            ["Дом","Квартира","Котельная","Промышленное здание",
             "Офис","Магазин","Склад","Другое"],
            cols=2, prefix="OT_"
        )
    )
    return U_OBJ_NEW_TYPE


async def u_obj_new_type(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    obj_type = q.data.replace("OT_", "")
    name     = ctx.user_data["_obj_name"]
    addr     = ctx.user_data.get("_obj_addr", "")
    user     = update.effective_user.full_name or ""
    with get_db() as db:
        db.execute(
            "INSERT INTO objects (name, address, obj_type, created_by, created_at) VALUES (?,?,?,?,?)",
            (name, addr, obj_type, user, datetime.now().strftime("%d.%m.%Y %H:%M"))
        )
        db.commit()
        obj_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    ctx.user_data["object_id"]   = obj_id
    ctx.user_data["object_name"] = name
    await q.edit_message_text(
        f"Объект {name} создан!\n\nВыбери бизнес-процесс:",
        reply_markup=bp_kb(),
    )
    return U_BP


# БП 
async def u_bp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    bp_id = int(q.data.split("_")[1])
    with get_db() as db:
        bp = db.execute("SELECT id, name FROM bp WHERE id=?", (bp_id,)).fetchone()
    ctx.user_data["bp_id"]   = bp["id"]
    ctx.user_data["bp_name"] = bp["name"]
    await q.edit_message_text(
        f"Объект: {ctx.user_data['object_name']}\n"
        f"БП: {bp['name']}\n\n"
        "Выбери задачу:",
        reply_markup=tasks_kb(bp_id),
    )
    return U_TASK


# Задача
async def u_task(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "TASK_CUSTOM":
        await q.edit_message_text("Напиши свою задачу:")
        return U_TASK_CUSTOM
    task_id = int(q.data.split("_")[1])
    with get_db() as db:
        task = db.execute("SELECT id, name FROM tasks_ref WHERE id=?", (task_id,)).fetchone()
    ctx.user_data["task_id"]   = task["id"]
    ctx.user_data["task_name"] = task["name"]

    # Если выбрана "Сдача работ" — объект будет закрыт после сохранения записи
    HANDOVER_KEYWORDS = ("сдача", "сдать", "передача", "закрыт", "завершение работ")
    if any(kw in task["name"].lower() for kw in HANDOVER_KEYWORDS):
        ctx.user_data["_close_object"] = True
        await q.edit_message_text(
            f"Задача: {task['name']}\n"
            "Объект будет помечен как СДАН после сохранения записи.\n\n"
            "Выбери процедуру:",
            reply_markup=procs_kb(task_id),
        )
    else:
        ctx.user_data["_close_object"] = False
        await q.edit_message_text(
            f"Задача: {task['name']}\n\nВыбери процедуру:",
            reply_markup=procs_kb(task_id),
        )
    return U_PROCEDURE


async def u_task_custom(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    ctx.user_data["task_id"]   = None
    ctx.user_data["task_name"] = name
    await update.message.reply_text(
        f"Задача: {name}\n\nНапиши процедуру (что конкретно делал):"
    )
    return U_PROC_CUSTOM


# Процедура 
async def u_procedure(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "PROC_CUSTOM":
        await q.edit_message_text("Напиши процедуру (что конкретно делал):")
        return U_PROC_CUSTOM
    proc_id = int(q.data.split("_")[1])
    with get_db() as db:
        proc = db.execute("SELECT id, name FROM procedures WHERE id=?", (proc_id,)).fetchone()
    ctx.user_data["proc_id"]   = proc["id"]
    ctx.user_data["proc_name"] = proc["name"]
    await q.edit_message_text(
        f"Процедура: {proc['name']}\n\n"
        "Введи время НАЧАЛА (ЧЧ:ММ), например: 09:30"
    )
    return U_TIME_START


async def u_proc_custom(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    ctx.user_data["proc_id"]   = None
    ctx.user_data["proc_name"] = name
    await update.message.reply_text(
        f"Процедура: {name}\n\n"
        "Введи время НАЧАЛА (ЧЧ:ММ), например: 09:30"
    )
    return U_TIME_START


# Время 
async def u_time_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        datetime.strptime(text, "%H:%M")
    except ValueError:
        await update.message.reply_text("Неверный формат. Пример: 09:30")
        return U_TIME_START
    ctx.user_data["time_start"] = text
    await update.message.reply_text("Введи время ОКОНЧАНИЯ (ЧЧ:ММ):")
    return U_TIME_END


async def u_time_end(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        datetime.strptime(text, "%H:%M")
    except ValueError:
        await update.message.reply_text("Неверный формат. Пример: 14:00")
        return U_TIME_END
    start = ctx.user_data.get("time_start","")
    h = dur_h(start, text)
    if h <= 0:
        await update.message.reply_text(
            f"Время {text} должно быть позже {start}. Введи ещё раз:"
        )
        return U_TIME_END
    ctx.user_data["time_end"] = text
    ctx.user_data["dur_h"]    = h
    await update.message.reply_text(
        f"Длительность: {h} ч\n\nРаботал один или с кем-то?",
        reply_markup=cw_kb(),
    )
    return U_COWORKERS


# Коллеги 
async def u_coworkers(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "CW_SOLO":
        ctx.user_data["coworkers"] = ""
        await q.edit_message_text(
            "Добавь описание если было что-то необычное\n"
            "или напиши - чтобы пропустить:"
        )
        return U_DESCRIPTION
    if q.data == "CW_CUSTOM":
        await q.edit_message_text("Введи имя коллеги:")
        return U_CW_INPUT
    name = q.data[3:]   # убираем "CW_"
    ctx.user_data["coworkers"] = name
    await q.edit_message_text(
        f"Коллега: {name}\n\n"
        "Добавь описание или напиши -:"
    )
    return U_DESCRIPTION


async def u_cw_input(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    ctx.user_data["coworkers"] = name
    await update.message.reply_text(
        f"Коллега: {name}\n\n"
        "Добавь описание или напиши -:"
    )
    return U_DESCRIPTION


# Описание и подтверждение
async def u_description(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    ctx.user_data["description"] = "" if text == "-" else text
    await update.message.reply_text(
        make_summary(ctx.user_data),
        reply_markup=confirm_kb(),
    )
    return U_CONFIRM


async def u_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "CNO":
        ctx.user_data.clear()
        await q.edit_message_text("Отменено. /start — новая запись.")
        return ConversationHandler.END
    rec_id = save_work(ctx.user_data)
    close_msg = ""
    if ctx.user_data.get("_close_object"):
        close_msg = f"\nОбъект {ctx.user_data.get('object_name','')} помечен как СДАН."
    await q.edit_message_text(
        f"Запись #{rec_id} сохранена!{close_msg}\n\n"
        "/start — добавить ещё задачу."
    )
    ctx.user_data.clear()
    return ConversationHandler.END


# АДМИН
def amenu_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Добавить объект",    callback_data="A_OBJ")],
        [InlineKeyboardButton("Список объектов",    callback_data="A_LOBJ")],
        [InlineKeyboardButton("Добавить БП",        callback_data="A_BP")],
        [InlineKeyboardButton("Добавить задачу",    callback_data="A_TASK")],
        [InlineKeyboardButton("Добавить процедуру", callback_data="A_PROC")],
        [InlineKeyboardButton("Справочник БП",      callback_data="A_LBP")],
        [InlineKeyboardButton("Выгрузить Excel",    callback_data="A_XLSX")],
        [InlineKeyboardButton("Шаблоны задач",   callback_data="A_TMPL_HINT")],
    ])


async def cmd_admin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Нет доступа.")
        return ConversationHandler.END
    with get_db() as db:
        nl = db.execute("SELECT COUNT(*) FROM work_log").fetchone()[0]
        no = db.execute("SELECT COUNT(*) FROM objects").fetchone()[0]
        nb = db.execute("SELECT COUNT(*) FROM bp WHERE active=1").fetchone()[0]
    await update.message.reply_text(
        f"Панель администратора\n\n"
        f"Записей в журнале: {nl}\n"
        f"Объектов:          {no}\n"
        f"Бизнес-процессов:  {nb}\n\n"
        "Выбери действие:",
        reply_markup=amenu_kb(),
    )
    return A_MENU


async def adm_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    d = q.data

    if d == "A_XLSX":
        await q.edit_message_text("Генерирую Excel...")
        path = export_to_excel()
        if path and path.exists():
            with open(path, "rb") as f:
                await q.message.reply_document(document=f, filename=path.name,
                                                caption="Выгрузка готова")
        else:
            await q.message.reply_text("Установи openpyxl: pip install openpyxl")
        await q.message.reply_text("Выбери действие:", reply_markup=amenu_kb())
        return A_MENU

    if d == "A_OBJ":
        await q.edit_message_text("Введи наименование нового объекта:")
        return A_OBJ_NAME

    if d == "A_LOBJ":
        with get_db() as db:
            objs = db.execute("SELECT id,name,address,obj_type,status FROM objects ORDER BY id").fetchall()
        lines = ["Список объектов:\n"]
        for o in objs:
            addr = f", {o['address']}" if o['address'] else ""
            lines.append(f"#{o['id']} {o['name']}{addr} [{o['obj_type']}] {o['status']}")
        await q.edit_message_text(
            "\n".join(lines)[:4000],
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад", callback_data="A_BACK")]])
        )
        return A_MENU

    if d == "A_BP":
        await q.edit_message_text("Введи название нового бизнес-процесса:")
        return A_BP_NAME

    if d == "A_TASK":
        with get_db() as db:
            bps = db.execute("SELECT id, name FROM bp WHERE active=1 ORDER BY name").fetchall()
        rows = [[InlineKeyboardButton(b["name"], callback_data=f"SBP_{b['id']}")] for b in bps]
        await q.edit_message_text("Выбери БП:", reply_markup=InlineKeyboardMarkup(rows))
        return A_SEL_BP

    if d == "A_PROC":
        with get_db() as db:
            bps = db.execute("SELECT id, name FROM bp WHERE active=1 ORDER BY name").fetchall()
        rows = [[InlineKeyboardButton(b["name"], callback_data=f"SBP_{b['id']}")] for b in bps]
        await q.edit_message_text(
            "Добавление процедуры\n\nШаг 1 — выбери БП:",
            reply_markup=InlineKeyboardMarkup(rows)
        )
        ctx.user_data["_proc_mode"] = True
        return A_SEL_BP

    if d == "A_LBP":
        with get_db() as db:
            rows = db.execute("""SELECT b.name, t.name as tn, p.name as pn
                FROM bp b JOIN tasks_ref t ON t.bp_id=b.id JOIN procedures p ON p.task_id=t.id
                WHERE b.active=1 AND t.active=1 AND p.active=1
                ORDER BY b.name, t.name, p.name""").fetchall()
        lines = []
        prev_b = prev_t = ""
        for r in rows:
            if r["name"] != prev_b:
                lines.append(f"\n=== {r['name']} ===")
                prev_b = r["name"]; prev_t = ""
            if r["tn"] != prev_t:
                lines.append(f"  [{r['tn']}]")
                prev_t = r["tn"]
            lines.append(f"    - {r['pn']}")
        await q.edit_message_text(
            "\n".join(lines)[:4000],
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад", callback_data="A_BACK")]])
        )
        return A_MENU

    if d == "A_TMPL_HINT":
        await q.edit_message_text(
            "Для управления шаблонами используй команду /templates\n\n"
            "Выбери действие:",
            reply_markup=amenu_kb()
        )
        return A_MENU

    if d == "A_BACK":
        await q.edit_message_text("Выбери действие:", reply_markup=amenu_kb())
        return A_MENU

    return A_MENU


async def adm_obj_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["_an"] = update.message.text.strip()
    await update.message.reply_text("Адрес объекта (или -):")
    return A_OBJ_ADDR


async def adm_obj_addr(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    addr = update.message.text.strip()
    ctx.user_data["_aa"] = "" if addr == "-" else addr
    await update.message.reply_text(
        "Тип объекта:",
        reply_markup=_kb(["Дом","Квартира","Котельная","Промышленное здание",
                           "Офис","Магазин","Склад","Другое"], cols=2, prefix="AT_")
    )
    return A_OBJ_TYPE


async def adm_obj_type(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    obj_type = q.data.replace("AT_", "")
    name = ctx.user_data["_an"]
    addr = ctx.user_data.get("_aa", "")
    user = update.effective_user.full_name or ""
    with get_db() as db:
        db.execute(
            "INSERT INTO objects (name, address, obj_type, created_by, created_at) VALUES (?,?,?,?,?)",
            (name, addr, obj_type, user, datetime.now().strftime("%d.%m.%Y %H:%M"))
        )
        db.commit()
    await q.edit_message_text(f"Объект {name} добавлен!\n\nВыбери действие:", reply_markup=amenu_kb())
    return A_MENU


async def adm_bp_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    with get_db() as db:
        if db.execute("SELECT id FROM bp WHERE name=?", (name,)).fetchone():
            await update.message.reply_text(f"БП «{name}» уже существует.")
            await update.message.reply_text("Выбери действие:", reply_markup=amenu_kb())
            return A_MENU
        db.execute("INSERT INTO bp (name) VALUES (?)", (name,))
        db.commit()
        bp_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    ctx.user_data["_tbp"] = bp_id
    ctx.user_data["_tbp_name"] = name
    await update.message.reply_text(
        f"БП «{name}» добавлен!\n\n"
        "Теперь добавь задачи этого БП.\n"
        "Введи название первой задачи\n"
        "(или - чтобы пропустить и закончить):"
    )
    return A_BP_TASK_NAME


async def adm_bp_task_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name  = update.message.text.strip()
    bp_id = ctx.user_data.get("_tbp")
    bp_nm = ctx.user_data.get("_tbp_name", "")
    if name == "-":
        await update.message.reply_text(
            f"Готово! БП «{bp_nm}» сохранён.\n"
            "Выбери действие:", reply_markup=amenu_kb()
        )
        return A_MENU
    with get_db() as db:
        db.execute("INSERT INTO tasks_ref (bp_id, name) VALUES (?,?)", (bp_id, name))
        db.commit()
    await update.message.reply_text(
        f"Задача «{name}» добавлена к БП «{bp_nm}».\n\n"
        "Введи следующую задачу или - чтобы закончить:"
    )
    return A_BP_TASK_NAME


async def adm_sel_bp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    bp_id = int(q.data.replace("SBP_", ""))
    ctx.user_data["_tbp"] = bp_id
    with get_db() as db:
        bp    = db.execute("SELECT name FROM bp WHERE id=?", (bp_id,)).fetchone()
        tasks = db.execute(
            "SELECT id, name FROM tasks_ref WHERE bp_id=? AND active=1 ORDER BY name",
            (bp_id,)
        ).fetchall()
    ctx.user_data["_tbp_name"] = bp["name"] if bp else ""

    if ctx.user_data.get("_proc_mode"):
        # Режим добавления процедуры — выбираем задачу
        rows = [[InlineKeyboardButton(t["name"], callback_data=f"ST_{t['id']}")] for t in tasks]
        await q.edit_message_text(
            f"БП: {bp['name']}\n\nШаг 2 — выбери задачу:",
            reply_markup=InlineKeyboardMarkup(rows)
        )
        return A_SEL_TASK
    else:
        # Режим добавления задачи — сразу вводим название
        await q.edit_message_text(f"БП: {bp['name']}\n\nВведи название задачи:")
        return A_TASK_NAME


async def adm_task_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name  = update.message.text.strip()
    bp_id = ctx.user_data["_tbp"]
    with get_db() as db:
        db.execute("INSERT INTO tasks_ref (bp_id, name) VALUES (?,?)", (bp_id, name))
        db.commit()
    await update.message.reply_text(f"Задача {name} добавлена!")
    await update.message.reply_text("Выбери действие:", reply_markup=amenu_kb())
    return A_MENU


async def adm_sel_task(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    task_id = int(q.data.replace("ST_", ""))
    ctx.user_data["_tt"] = task_id
    with get_db() as db:
        task = db.execute("SELECT name FROM tasks_ref WHERE id=?", (task_id,)).fetchone()
    await q.edit_message_text(f"Задача: {task['name']}\n\nВведи название процедуры:")
    return A_PROC_NAME


async def adm_proc_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name    = update.message.text.strip()
    task_id = ctx.user_data["_tt"]
    with get_db() as db:
        db.execute("INSERT INTO procedures (task_id, name) VALUES (?,?)", (task_id, name))
        db.commit()
    ctx.user_data.pop("_proc_mode", None)  # сбрасываем режим
    ctx.user_data.pop("_tt", None)
    ctx.user_data.pop("_tbp", None)
    ctx.user_data.pop("_tbp_name", None)
    await update.message.reply_text(f"Процедура «{name}» добавлена!")
    await update.message.reply_text("Выбери действие:", reply_markup=amenu_kb())
    return A_MENU


# СБОРКА

# ПЛАНИРОВЩИК — ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
def plan_status_kb(task_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Выполнена",    callback_data=f"PS_done_{task_id}")],
        [InlineKeyboardButton("🔄 В процессе",  callback_data=f"PS_wip_{task_id}")],
        [InlineKeyboardButton("❌ Не выполнена", callback_data=f"PS_fail_{task_id}")],
        [InlineKeyboardButton("ℹ️ Подробнее",   callback_data=f"PS_info_{task_id}")],
    ])


def fail_reason_kb(task_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔁 Перенести на другой день", callback_data=f"PF_move_{task_id}")],
        [InlineKeyboardButton("✏️ Нужно исправить",          callback_data=f"PF_fix_{task_id}")],
    ])


def _date_from_text(text: str):
    from datetime import timedelta
    t = text.strip().lower()
    if t in ("сегодня", "today"):
        return datetime.now().strftime("%d.%m.%Y")
    if t in ("завтра", "tomorrow"):
        return (datetime.now() + timedelta(days=1)).strftime("%d.%m.%Y")
    try:
        return datetime.strptime(t, "%d.%m.%Y").strftime("%d.%m.%Y")
    except ValueError:
        return None


def _is_past_date(date_str: str) -> bool:
    try:
        d = datetime.strptime(date_str, "%d.%m.%Y")
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        return d < today
    except ValueError:
        return False


def save_plan(d: dict) -> int:
    with get_db() as db:
        db.execute("""
            INSERT INTO planned_tasks
                (title, object_name, assignee_name, assignee_tg_id,
                 planned_date, planned_time, planned_time_end,
                 bp_name, task_ref_name,
                 status, day_status, consistency,
                 created_by, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            d["title"],
            d.get("object_name", ""),
            d["assignee_name"],
            d.get("assignee_tg_id", ""),
            d["planned_date"],
            d.get("planned_time", ""),
            d.get("planned_time_end", ""),
            d.get("bp_name", ""),
            d.get("task_ref_name", ""),
            "Запланирована", "", "Согласована",
            d.get("created_by", ""),
            datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        ))
        db.commit()
        return db.execute("SELECT last_insert_rowid()").fetchone()[0]


def fmt_plan(row) -> str:
    icon = {"Запланирована":"🔵","Выполнена":"✅","В процессе":"🔄","Не выполнена":"❌"}.get(row["status"],"•")
    time_part = f"  {row['planned_time']}" if row["planned_time"] else ""
    obj_part  = f"\n   Объект: {row['object_name']}" if row["object_name"] else ""
    return (
        f"{icon} #{row['id']} {row['title']}\n"
        f"   Исполнитель: {row['assignee_name']}{time_part}"
        f"{obj_part}"
    )


def get_today_plans():
    today = datetime.now().strftime("%d.%m.%Y")
    with get_db() as db:
        return db.execute(
            "SELECT * FROM planned_tasks WHERE planned_date=? ORDER BY planned_time, id",
            (today,)
        ).fetchall()


def get_week_plans():
    from datetime import timedelta
    today = datetime.now()
    days  = [(today + timedelta(days=i)).strftime("%d.%m.%Y") for i in range(7)]
    ph    = ",".join("?" * 7)
    with get_db() as db:
        return db.execute(
            f"SELECT * FROM planned_tasks WHERE planned_date IN ({ph}) ORDER BY planned_date, planned_time",
            days
        ).fetchall()


# Напоминания 

async def _remind_job(context: ContextTypes.DEFAULT_TYPE):
    now   = datetime.now()
    today = now.strftime("%d.%m.%Y")

    with get_db() as db:
        tasks = db.execute(
            "SELECT * FROM planned_tasks WHERE planned_date=? AND status='Запланирована'",
            (today,)
        ).fetchall()

    for task in tasks:
        if not task["planned_time"] or not task["assignee_tg_id"]:
            continue
        try:
            pt = datetime.strptime(f"{task['planned_date']} {task['planned_time']}", "%d.%m.%Y %H:%M")
        except ValueError:
            continue

        diff_min = (pt - now).total_seconds() / 60

        # Напоминание за 2 часа
        if 115 <= diff_min <= 125 and not task["reminded_2h"]:
            try:
                await context.bot.send_message(
                    chat_id=int(task["assignee_tg_id"]),
                    text=(
                        f"Напоминание — через 2 часа:\n\n"
                        f"Задача: {task['title']}\n"
                        f"Время:  {task['planned_time']}\n"
                        f"Объект: {task['object_name'] or 'не указан'}\n\n"
                        "Отметить выполнение: /tasks"
                    ),
                    reply_markup=plan_status_kb(task["id"]),
                )
                with get_db() as db:
                    db.execute("UPDATE planned_tasks SET reminded_2h=1 WHERE id=?", (task["id"],))
                    db.commit()
            except Exception as e:
                log.warning(f"Напоминание 2ч не отправлено ({task['id']}): {e}")

    # Утреннее напоминание в 07:00
    if now.hour == 7 and now.minute < 10:
        with get_db() as db:
            day_tasks = db.execute(
                "SELECT * FROM planned_tasks WHERE planned_date=? AND reminded_day=0",
                (today,)
            ).fetchall()

        by_user: dict = {}
        for t in day_tasks:
            by_user.setdefault(t["assignee_tg_id"], []).append(t)

        for uid, utasks in by_user.items():
            if not uid:
                continue
            lines = [f"Доброе утро! Задачи на сегодня ({today}):\n"]
            for t in utasks:
                tp = f" в {t['planned_time']}" if t["planned_time"] else ""
                lines.append(f"  {t['title']}{tp}")
                if t["object_name"]:
                    lines.append(f"    Объект: {t['object_name']}")
            lines.append("\nОтметить выполнение: /tasks")
            try:
                await context.bot.send_message(chat_id=int(uid), text="\n".join(lines))
                with get_db() as db:
                    for t in utasks:
                        db.execute("UPDATE planned_tasks SET reminded_day=1 WHERE id=?", (t["id"],))
                    db.commit()
            except Exception as e:
                log.warning(f"Утреннее напоминание не отправлено ({uid}): {e}")


def _start_scheduler(app: Application):
    app.job_queue.run_repeating(_remind_job, interval=300, first=10)


# Диалог /plan 
async def cmd_plan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["_p"] = {}
    # Шаг 1 — сначала ДАТА (не в прошлом)
    from datetime import timedelta
    today = datetime.now().strftime("%d.%m.%Y")
    tmr   = (datetime.now() + timedelta(days=1)).strftime("%d.%m.%Y")
    await update.message.reply_text(
        "Новая задача в расписание\n\n"
        "Шаг 1 из 8 — Выбери дату\n"
        f"Сегодня: {today}\n"
        "Введи: сегодня / завтра / ДД.ММ.ГГГГ\n"
        "(Дата в прошлом не принимается)"
    )
    return P_DATE


async def p_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    date_str = _date_from_text(update.message.text)
    if not date_str:
        await update.message.reply_text("Неверный формат. Введи: сегодня / завтра / ДД.ММ.ГГГГ")
        return P_DATE
    if _is_past_date(date_str):
        await update.message.reply_text(
            f"Дата {date_str} уже прошла. Введи сегодня или будущую дату:"
        )
        return P_DATE
    ctx.user_data["_p"]["planned_date"] = date_str
    await update.message.reply_text(
        f"Дата: {date_str}\n\n"
        "Шаг 2 из 8 — Название задачи\n"
        "Например: Монтаж котла, Встреча с клиентом Ивановым"
    )
    return P_TITLE


async def p_title(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    title = update.message.text.strip()
    if not title:
        await update.message.reply_text("Название не может быть пустым. Введи ещё раз:")
        return P_TITLE
    ctx.user_data["_p"]["title"] = title
    with get_db() as db:
        objs = db.execute(
            "SELECT id, name FROM objects WHERE status != 'Сдан' ORDER BY name"
        ).fetchall()
    rows = [[InlineKeyboardButton(f"📍 {o['name']}", callback_data=f"PO_{o['id']}")] for o in objs]
    rows.append([InlineKeyboardButton("Без объекта / пропустить", callback_data="PO_SKIP")])
    await update.message.reply_text(
        f"Задача: {title}\n\n"
        "Шаг 3 из 8 — Объект:",
        reply_markup=InlineKeyboardMarkup(rows),
    )
    return P_OBJECT


async def p_object(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "PO_SKIP":
        ctx.user_data["_p"]["object_name"] = ""
        ctx.user_data["_p"]["object_id"]   = None
    else:
        obj_id = int(q.data.replace("PO_", ""))
        with get_db() as db:
            obj = db.execute("SELECT id, name FROM objects WHERE id=?", (obj_id,)).fetchone()
        ctx.user_data["_p"]["object_name"] = obj["name"] if obj else ""
        ctx.user_data["_p"]["object_id"]   = obj_id
    obj_txt = ctx.user_data["_p"]["object_name"] or "не указан"
    with get_db() as db:
        bps = db.execute("SELECT id, name FROM bp WHERE active=1 ORDER BY name").fetchall()
    rows = [[InlineKeyboardButton(b["name"], callback_data=f"PP_BP_{b['id']}")] for b in bps]
    rows.append([InlineKeyboardButton("Пропустить БП", callback_data="PP_BP_SKIP")])
    await q.edit_message_text(
        f"Объект: {obj_txt}\n\n"
        "Шаг 4 из 8 — Бизнес-процесс:",
        reply_markup=InlineKeyboardMarkup(rows),
    )
    return P_BP


async def p_bp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "PP_BP_SKIP":
        ctx.user_data["_p"]["bp_name"]       = ""
        ctx.user_data["_p"]["bp_id"]         = None
        ctx.user_data["_p"]["task_ref_name"] = ""
        await q.edit_message_text(
            "Шаг 5 из 8 — Задача\nВведи название задачи вручную:"
        )
        return P_TASK
    bp_id = int(q.data.replace("PP_BP_", ""))
    with get_db() as db:
        bp    = db.execute("SELECT name FROM bp WHERE id=?", (bp_id,)).fetchone()
        tasks = db.execute(
            "SELECT id, name FROM tasks_ref WHERE bp_id=? AND active=1 ORDER BY name",
            (bp_id,)
        ).fetchall()
    ctx.user_data["_p"]["bp_name"] = bp["name"] if bp else ""
    ctx.user_data["_p"]["bp_id"]   = bp_id
    rows = [[InlineKeyboardButton(t["name"], callback_data=f"PP_T_{t['id']}")] for t in tasks]
    rows.append([InlineKeyboardButton("✏️ Ввести вручную", callback_data="PP_T_CUSTOM")])
    await q.edit_message_text(
        f"БП: {ctx.user_data['_p']['bp_name']}\n\n"
        "Шаг 5 из 8 — Задача:",
        reply_markup=InlineKeyboardMarkup(rows),
    )
    return P_TASK


async def p_task(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "PP_T_CUSTOM":
        await q.edit_message_text("Введи название задачи:")
        return P_TASK
    task_id = int(q.data.replace("PP_T_", ""))
    with get_db() as db:
        t = db.execute("SELECT name FROM tasks_ref WHERE id=?", (task_id,)).fetchone()
    ctx.user_data["_p"]["task_ref_name"] = t["name"] if t else ""
    return await _p_to_assignee(update.callback_query, ctx)


async def p_task_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["_p"]["task_ref_name"] = update.message.text.strip()
    return await _p_to_assignee(update.message, ctx)


async def _p_to_assignee(src_obj, ctx):
    task_name = (ctx.user_data["_p"].get("task_ref_name") or
                 ctx.user_data["_p"].get("title", ""))
    with get_db() as db:
        names = db.execute(
            "SELECT DISTINCT employee_name FROM work_log ORDER BY employee_name LIMIT 8"
        ).fetchall()
    rows = [[InlineKeyboardButton(n["employee_name"], callback_data=f"PA_{n['employee_name']}")] for n in names]
    rows.append([InlineKeyboardButton("✏️ Ввести имя вручную", callback_data="PA_CUSTOM")])
    text = f"Задача: {task_name}\n\nШаг 6 из 8 — Исполнитель:"
    if hasattr(src_obj, "edit_message_text"):
        await src_obj.edit_message_text(text, reply_markup=InlineKeyboardMarkup(rows))
    else:
        await src_obj.reply_text(text, reply_markup=InlineKeyboardMarkup(rows))
    return P_ASSIGNEE


async def p_assignee_btn(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "PA_CUSTOM":
        await q.edit_message_text("Введи имя исполнителя:")
        return P_ASSIGNEE
    name = q.data[3:]
    ctx.user_data["_p"]["assignee_name"] = name
    with get_db() as db:
        row = db.execute(
            "SELECT employee_tg_id FROM work_log WHERE employee_name=? LIMIT 1", (name,)
        ).fetchone()
    ctx.user_data["_p"]["assignee_tg_id"] = row["employee_tg_id"] if row else ""
    await q.edit_message_text(
        f"Исполнитель: {name}\n\n"
        "Шаг 7 из 8 — Время начала (ЧЧ:ММ)\n"
        "Или - чтобы пропустить:"
    )
    return P_TIME_START


async def p_assignee_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    ctx.user_data["_p"]["assignee_name"]  = name
    ctx.user_data["_p"]["assignee_tg_id"] = ""
    await update.message.reply_text(
        f"Исполнитель: {name}\n\n"
        "Шаг 7 из 8 — Время начала (ЧЧ:ММ)\n"
        "Или - чтобы пропустить:"
    )
    return P_TIME_START


async def p_time_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text == "-":
        ctx.user_data["_p"]["planned_time"]     = ""
        ctx.user_data["_p"]["planned_time_end"] = ""
        return await _show_plan_confirm(update.message, ctx)
    try:
        datetime.strptime(text, "%H:%M")
        ctx.user_data["_p"]["planned_time"] = text
    except ValueError:
        await update.message.reply_text("Неверный формат. Пример: 09:30 или - пропустить:")
        return P_TIME_START
    await update.message.reply_text(
        f"Начало: {text}\n\n"
        "Шаг 8 из 8 — Примерное время окончания (ЧЧ:ММ)\n"
        "Или - пропустить:"
    )
    return P_TIME_END


async def p_time_end(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text == "-":
        ctx.user_data["_p"]["planned_time_end"] = ""
    else:
        try:
            datetime.strptime(text, "%H:%M")
            ctx.user_data["_p"]["planned_time_end"] = text
        except ValueError:
            await update.message.reply_text("Неверный формат. Пример: 11:00 или - пропустить:")
            return P_TIME_END
    return await _show_plan_confirm(update.message, ctx)


async def _show_plan_confirm(msg, ctx):
    p = ctx.user_data["_p"]
    lines = ["Проверь задачу:\n",
             f"Дата:        {p['planned_date']}",
             f"Название:    {p['title']}"]
    if p.get("bp_name"):
        lines.append(f"БП:          {p['bp_name']}")
    if p.get("task_ref_name"):
        lines.append(f"Задача:      {p['task_ref_name']}")
    if p.get("object_name"):
        lines.append(f"Объект:      {p['object_name']}")
    lines.append(f"Исполнитель: {p['assignee_name']}")
    if p.get("planned_time"):
        t_line = f"Начало:      {p['planned_time']}"
        if p.get("planned_time_end"):
            t_line += f"  →  {p['planned_time_end']}"
        lines.append(t_line)
    await msg.reply_text(
        "\n".join(lines),
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Сохранить", callback_data="PC_YES"),
            InlineKeyboardButton("❌ Отменить",  callback_data="PC_NO"),
        ]])
    )
    return P_CONFIRM


async def p_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "PC_NO":
        ctx.user_data.pop("_p", None)
        await q.edit_message_text("Отменено. /plan — новая задача, /tasks — список.")
        return ConversationHandler.END
    p = ctx.user_data["_p"]
    p["created_by"] = update.effective_user.full_name or ""
    rec_id = save_plan(p)
    await q.edit_message_text(
        f"Задача #{rec_id} добавлена!\n\n"
        f"Дата: {p['planned_date']}\n"
        "/tasks — задачи на сегодня\n"
        "/week  — задачи на неделю\n"
        "/plan  — добавить ещё"
    )
    ctx.user_data.pop("_p", None)
    return ConversationHandler.END


# Просмотр задач 
async def cmd_tasks(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    today = datetime.now().strftime("%d.%m.%Y")
    rows  = get_today_plans()
    if not rows:
        await update.message.reply_text(
            f"На сегодня ({today}) задач нет.\n/plan — добавить задачу"
        )
        return
    total   = len(rows)
    done    = sum(1 for r in rows if r["status"] == "Выполнена")
    pending = sum(1 for r in rows if r["status"] == "Запланирована")
    failed  = sum(1 for r in rows if r["status"] == "Не выполнена")
    await update.message.reply_text(
        f"Задачи на сегодня ({today})\n"
        f"Всего: {total}  Выполнено: {done}  "
        f"Ожидает: {pending}  Не выполнено: {failed}"
    )
    for row in rows:
        await update.message.reply_text(fmt_plan(row), reply_markup=plan_status_kb(row["id"]))


async def cmd_week(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    rows = get_week_plans()
    if not rows:
        await update.message.reply_text("На ближайшую неделю задач нет.\n/plan — добавить задачу")
        return
    by_date: dict = {}
    for r in rows:
        by_date.setdefault(r["planned_date"], []).append(r)
    today = datetime.now().strftime("%d.%m.%Y")
    icons = {"Запланирована":"🔵","Выполнена":"✅","В процессе":"🔄","Не выполнена":"❌"}
    lines = ["Задачи на неделю:\n"]
    for date, tasks in by_date.items():
        mark = " (сегодня)" if date == today else ""
        lines.append(f"\n{date}{mark}")
        for t in tasks:
            tp = f" {t['planned_time']}" if t["planned_time"] else ""
            lines.append(f"  {icons.get(t['status'],'•')} {t['title']}{tp} — {t['assignee_name']}")
    lines.append("\n/tasks — подробно на сегодня с кнопками")
    await update.message.reply_text("\n".join(lines)[:4096])


# Обновление статуса 
async def plan_status_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    parts   = q.data.split("_")
    prefix  = parts[0]  # PS или PF
    action  = parts[1]
    task_id = int(parts[2])

    # Подменю причины невыполнения (PF_)
    if prefix == "PF":
        with get_db() as db:
            row = db.execute("SELECT * FROM planned_tasks WHERE id=?", (task_id,)).fetchone()
        if not row:
            return
        if action == "move":
            ctx.user_data["_reschedule_id"] = task_id
            await q.edit_message_text(
                f"Задача «{row['title']}» — перенос\n\n"
                "Введи новую дату (сегодня / завтра / ДД.ММ.ГГГГ):"
            )
            ctx.user_data["_await_reschedule"] = True
            return
        if action == "fix":
            ctx.user_data["_fail_fix_id"] = task_id
            await q.edit_message_text(
                f"Задача «{row['title']}» — нужно исправить\n\n"
                "Опиши кратко причину, почему не удалось выполнить:"
            )
            ctx.user_data["_await_fail_reason"] = True
            return
        return

    # Основные кнопки статуса (PS_) 
    if action == "info":
        with get_db() as db:
            row = db.execute("SELECT * FROM planned_tasks WHERE id=?", (task_id,)).fetchone()
        if row:
            tp  = f"\nВремя:       {row['planned_time']}" if row["planned_time"] else ""
            op  = f"\nОбъект:      {row['object_name']}"  if row["object_name"]  else ""
            bp  = f"\nБП:          {row['bp_name']}"       if row.get("bp_name")  else ""
            tsk = f"\nЗадача:      {row['task_ref_name']}" if row.get("task_ref_name") else ""
            fr  = f"\nПричина:     {row['fail_reason']}"   if row.get("fail_reason") else ""
            rsc = f"\nПеренесена:  {row['rescheduled_to']}" if row.get("rescheduled_to") else ""
            await q.edit_message_text(
                f"Задача #{row['id']}\n"
                f"Название:    {row['title']}"
                f"{bp}{tsk}{op}"
                f"\nИсполнитель: {row['assignee_name']}"
                f"\nДата:        {row['planned_date']}"
                f"{tp}"
                f"\nСтатус:      {row['status']}"
                f"{fr}{rsc}",
                reply_markup=plan_status_kb(task_id)
            )
        return

    if action == "done":
        with get_db() as db:
            db.execute(
                "UPDATE planned_tasks SET status='Выполнена', day_status='Выполнена' WHERE id=?",
                (task_id,)
            )
            db.commit()
            row = db.execute("SELECT title FROM planned_tasks WHERE id=?", (task_id,)).fetchone()
        await q.edit_message_text(f"✅ #{task_id} {row['title']}\nСтатус: Выполнена")
        return

    if action == "wip":
        with get_db() as db:
            db.execute(
                "UPDATE planned_tasks SET status='В процессе', day_status='В процессе' WHERE id=?",
                (task_id,)
            )
            db.commit()
            row = db.execute("SELECT title FROM planned_tasks WHERE id=?", (task_id,)).fetchone()
        await q.edit_message_text(
            f"🔄 #{task_id} {row['title']}\nСтатус: В процессе",
            reply_markup=plan_status_kb(task_id)
        )
        return

    if action == "fail":
        with get_db() as db:
            db.execute(
                "UPDATE planned_tasks SET status='Не выполнена', day_status='Не выполнена' WHERE id=?",
                (task_id,)
            )
            db.commit()
            row = db.execute("SELECT title FROM planned_tasks WHERE id=?", (task_id,)).fetchone()
        await q.edit_message_text(
            f"❌ #{task_id} {row['title']}\nСтатус: Не выполнена\n\nЧто делаем?",
            reply_markup=fail_reason_kb(task_id)
        )
        return


async def handle_reschedule_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    # Ранний выход если нет активных флагов ожидания
    if not ctx.user_data.get("_await_reschedule") and not ctx.user_data.get("_await_fail_reason"):
        return
    """Обрабатывает текстовый ввод для переноса / причины."""
    if ctx.user_data.get("_await_reschedule"):
        task_id  = ctx.user_data.pop("_reschedule_id", None)
        ctx.user_data.pop("_await_reschedule", None)
        date_str = _date_from_text(update.message.text)
        if not date_str:
            await update.message.reply_text(
                "Неверный формат. Введи: сегодня / завтра / ДД.ММ.ГГГГ"
            )
            ctx.user_data["_await_reschedule"] = True
            ctx.user_data["_reschedule_id"]    = task_id
            return
        if _is_past_date(date_str):
            await update.message.reply_text(
                f"Дата {date_str} уже прошла. Введи будущую дату:"
            )
            ctx.user_data["_await_reschedule"] = True
            ctx.user_data["_reschedule_id"]    = task_id
            return
        with get_db() as db:
            row = db.execute("SELECT * FROM planned_tasks WHERE id=?", (task_id,)).fetchone()
        if not row:
            await update.message.reply_text("Задача не найдена.")
            return
        # Создаём новую запись на новую дату
        new_d = {
            "title":          row["title"],
            "object_name":    row["object_name"],
            "assignee_name":  row["assignee_name"],
            "assignee_tg_id": row["assignee_tg_id"],
            "planned_date":   date_str,
            "planned_time":   row["planned_time"],
            "bp_name":        row.get("bp_name", ""),
            "task_ref_name":  row.get("task_ref_name", ""),
            "created_by":     update.effective_user.full_name or "",
        }
        new_id = save_plan(new_d)
        with get_db() as db:
            db.execute(
                "UPDATE planned_tasks SET rescheduled_to=? WHERE id=?",
                (date_str, task_id)
            )
            db.commit()
        await update.message.reply_text(
            f"Задача перенесена на {date_str}.\n"
            f"Создана новая запись #{new_id}.\n"
            "/week — посмотреть расписание"
        )
        return

    if ctx.user_data.get("_await_fail_reason"):
        task_id = ctx.user_data.pop("_fail_fix_id", None)
        ctx.user_data.pop("_await_fail_reason", None)
        reason  = update.message.text.strip()
        with get_db() as db:
            db.execute(
                "UPDATE planned_tasks SET fail_reason=? WHERE id=?",
                (reason, task_id)
            )
            db.commit()
            row = db.execute("SELECT title FROM planned_tasks WHERE id=?", (task_id,)).fetchone()
        await update.message.reply_text(
            f"Причина зафиксирована для задачи «{row['title'] if row else task_id}»:\n"
            f"{reason}\n\n"
            "Задача помечена как требующая исправления."
        )
        return


# Статистика для администратора 
async def cmd_stats(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Нет доступа.")
        return
    today = datetime.now().strftime("%d.%m.%Y")
    with get_db() as db:
        by_status = {r["status"]: r["cnt"] for r in db.execute(
            "SELECT status, COUNT(*) as cnt FROM planned_tasks WHERE planned_date=? GROUP BY status",
            (today,)
        ).fetchall()}
        total_all  = db.execute("SELECT COUNT(*) FROM planned_tasks").fetchone()[0]
        done_all   = db.execute("SELECT COUNT(*) FROM planned_tasks WHERE status='Выполнена'").fetchone()[0]
        top        = db.execute("""
            SELECT assignee_name, COUNT(*) as cnt,
                   SUM(CASE WHEN status='Выполнена' THEN 1 ELSE 0 END) as done
            FROM planned_tasks WHERE planned_date=?
            GROUP BY assignee_name ORDER BY cnt DESC
        """, (today,)).fetchall()
        # Выполненные задачи из /start за сегодня
        work_today = db.execute("""
            SELECT employee_name, COUNT(*) as cnt, SUM(duration_h) as total_h
            FROM work_log WHERE date=?
            GROUP BY employee_name ORDER BY cnt DESC
        """, (today,)).fetchall()
        work_total = db.execute("SELECT COUNT(*) FROM work_log WHERE date=?", (today,)).fetchone()[0]

    lines = [
        f"Статистика задач ({today})\n",
        "— Расписание (planned_tasks) —",
        f"  Запланировано:  {by_status.get('Запланирована', 0)}",
        f"  В процессе:     {by_status.get('В процессе', 0)}",
        f"  Выполнено:      {by_status.get('Выполнена', 0)}",
        f"  Не выполнено:   {by_status.get('Не выполнена', 0)}",
        f"\n— Зафиксированных задач /start за сегодня: {work_total} —",
    ]
    if work_today:
        for w in work_today:
            h = round(w["total_h"] or 0, 2)
            lines.append(f"  {w['employee_name']}: {w['cnt']} задач  ({h} ч)")
    if top:
        lines.append("\nПо расписанию — исполнители:")
        for t in top:
            lines.append(f"  {t['assignee_name']}: {t['cnt']} задач, выполнено {t['done']}")
    lines.append(f"\nВсего в базе planned: {total_all}, выполнено: {done_all}")
    lines.append("\n/tasks — задачи на сегодня\n/week — задачи на неделю")
    await update.message.reply_text("\n".join(lines))

async def admin_templates_cmd(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Нет доступа.")
        return
    # Импортируем и вызываем меню шаблонов
    from templates_module import admin_templates_menu
    await admin_templates_menu(update, ctx)

def build_app() -> Application:
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()

    user_conv = ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
            U_OBJECT:       [CallbackQueryHandler(u_object)],
            U_OBJ_NEW_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, u_obj_new_name)],
            U_OBJ_NEW_ADDR: [MessageHandler(filters.TEXT & ~filters.COMMAND, u_obj_new_addr)],
            U_OBJ_NEW_TYPE: [CallbackQueryHandler(u_obj_new_type, pattern="^OT_")],
            U_BP:           [CallbackQueryHandler(u_bp,  pattern="^BP_")],
            U_TASK:         [CallbackQueryHandler(u_task, pattern="^TASK_")],
            U_TASK_CUSTOM:  [MessageHandler(filters.TEXT & ~filters.COMMAND, u_task_custom)],
            U_PROCEDURE:    [CallbackQueryHandler(u_procedure, pattern="^PROC_")],
            U_PROC_CUSTOM:  [MessageHandler(filters.TEXT & ~filters.COMMAND, u_proc_custom)],
            U_TIME_START:   [MessageHandler(filters.TEXT & ~filters.COMMAND, u_time_start)],
            U_TIME_END:     [MessageHandler(filters.TEXT & ~filters.COMMAND, u_time_end)],
            U_COWORKERS:    [CallbackQueryHandler(u_coworkers, pattern="^CW_")],
            U_CW_INPUT:     [MessageHandler(filters.TEXT & ~filters.COMMAND, u_cw_input)],
            U_DESCRIPTION:  [MessageHandler(filters.TEXT & ~filters.COMMAND, u_description)],
            U_CONFIRM:      [CallbackQueryHandler(u_confirm, pattern="^C")],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
        allow_reentry=True,
    )

    admin_conv = ConversationHandler(
        entry_points=[CommandHandler("admin", cmd_admin)],
        states={
            A_MENU:         [CallbackQueryHandler(adm_menu)],
            A_OBJ_NAME:     [MessageHandler(filters.TEXT & ~filters.COMMAND, adm_obj_name)],
            A_OBJ_ADDR:     [MessageHandler(filters.TEXT & ~filters.COMMAND, adm_obj_addr)],
            A_OBJ_TYPE:     [CallbackQueryHandler(adm_obj_type, pattern="^AT_")],
            A_BP_NAME:      [MessageHandler(filters.TEXT & ~filters.COMMAND, adm_bp_name)],
            A_BP_TASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, adm_bp_task_name)],
            A_SEL_BP:       [CallbackQueryHandler(adm_sel_bp,   pattern="^SBP_")],
            A_TASK_NAME:    [MessageHandler(filters.TEXT & ~filters.COMMAND, adm_task_name)],
            A_SEL_TASK:     [CallbackQueryHandler(adm_sel_task, pattern="^ST_")],
            A_PROC_NAME:    [MessageHandler(filters.TEXT & ~filters.COMMAND, adm_proc_name)],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
        allow_reentry=True,
    )

    # Планировщик задач на день/неделю
    plan_conv = ConversationHandler(
        entry_points=[CommandHandler("plan", cmd_plan)],
        states={
            P_DATE:       [MessageHandler(filters.TEXT & ~filters.COMMAND, p_date)],
            P_TITLE:      [MessageHandler(filters.TEXT & ~filters.COMMAND, p_title)],
            P_OBJECT:     [CallbackQueryHandler(p_object, pattern="^PO_")],
            P_BP:         [CallbackQueryHandler(p_bp,     pattern="^PP_BP_")],
            P_TASK: [
                CallbackQueryHandler(p_task,     pattern="^PP_T_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, p_task_text),
            ],
            P_ASSIGNEE: [
                CallbackQueryHandler(p_assignee_btn, pattern="^PA_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, p_assignee_text),
            ],
            P_TIME_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, p_time_start)],
            P_TIME_END:   [MessageHandler(filters.TEXT & ~filters.COMMAND, p_time_end)],
            P_CONFIRM:    [CallbackQueryHandler(p_confirm, pattern="^PC_")],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("templates", admin_templates_cmd))

    _start_scheduler(app)
    app.add_handler(user_conv)
    app.add_handler(admin_conv)
    app.add_handler(plan_conv)
    app.add_handler(CommandHandler("tasks", cmd_tasks))
    app.add_handler(CommandHandler("week",  cmd_week))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("help",  cmd_help))
    app.add_handler(CommandHandler("info",  cmd_info))
    app.add_handler(CallbackQueryHandler(plan_status_handler, pattern="^PS_"))
    app.add_handler(CallbackQueryHandler(plan_status_handler, pattern="^PF_"))
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND, handle_reschedule_text
    ))

    for h in get_template_handlers():
        app.add_handler(h)
    for h in get_admin_template_handlers():
        app.add_handler(h)

    return app


if __name__ == "__main__":
    log.info("Бот запускается...")
    build_app().run_polling(drop_pending_updates=True)